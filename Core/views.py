from django.shortcuts import render,redirect
from django.contrib.auth.decorators import login_required
from . models import *
from datetime import datetime, date
from django.http import HttpResponse
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from datetime import datetime, timedelta
from django.db import IntegrityError
from django.contrib import messages

# Create your views here.


@login_required
def index(request):
    employees_count = Employees.objects.count()
    project_count = Projects.objects.count()
    # Get the current user
    employee = request.user

    # Get the current date
    today = datetime.now().date()

    # Filter timesheets for the current user and current date
    timesheets = TimeSheet.objects.filter(employee=employee, date=today)

    # Calculate the total duration for today
    total_duration = timedelta()
    for timesheet in timesheets:
        start_datetime = datetime.combine(today, timesheet.start_time)
        end_datetime = datetime.combine(today, timesheet.end_time)
        duration = end_datetime - start_datetime
        total_duration += duration

    # Convert the total duration to string format
    hours, remainder = divmod(total_duration.total_seconds(), 3600)
    minutes, seconds = divmod(remainder, 60)
    total_duration_str = f"{int(hours):02d}:{int(minutes):02d}:{int(seconds):02d}"

    return render(request, 'index.html', {'total_duration': total_duration_str,
                                          'employees_count': employees_count,
                                          'project_count': project_count})


# ------------------------ Projects ----------------------- #

@login_required
def project_list(request):
    projects = Projects.objects.all()
    return render(request, 'project_list.html', {'projects': projects})

@login_required
def project_create(request):
    if request.method == 'POST':
        project_name = request.POST.get('project_name')
        project_client_name = request.POST.get('project_client_name')
        project_client_number = request.POST.get('project_client_number')
        project_client_address = request.POST.get('project_client_address')

        Projects.objects.create(
            project_name=project_name,
            project_client_name=project_client_name,
            project_client_number=project_client_number,
            project_client_address=project_client_address
        )
        return redirect('project-list')

    return render(request, 'project_create.html')

@login_required
def project_update(request, pk):
    project = Projects.objects.get(pk=pk)

    if request.method == 'POST':
        project_name = request.POST.get('project_name')
        project_client_name = request.POST.get('project_client_name')
        project_client_number = request.POST.get('project_client_number')
        project_client_address = request.POST.get('project_client_address')

        project.project_name = project_name
        project.project_client_name = project_client_name
        project.project_client_number = project_client_number
        project.project_client_address = project_client_address
        project.save()
        return redirect('project-list')

    return render(request, 'project_update.html', {'project': project})

@login_required
def project_delete(request, pk):
    project = Projects.objects.get(pk=pk)
    project.delete()
    return redirect('project-list')



# ---------------------------- Employees ------------------------------------- #

@login_required
def employee_list(request):
    employees = Employees.objects.all()
    return render(request, 'employee_list.html', {'employees': employees})


def employee_details(request, employee_id):
    employee = Employees.objects.get(id=employee_id)
    return render(request, 'employee_details.html', {'employee': employee})


# @login_required
# def employee_create(request):
#     if request.method == 'POST':
#         username = request.POST.get('username')
#         password = request.POST.get('password')
#         first_name = request.POST.get('first_name')
#         last_name = request.POST.get('last_name')
#         employees_photo = request.FILES.get('employees_photo')
#         employees_job = request.POST.get('employees_job')
#         employees_number = request.POST.get('employees_number')
#         employees_address = request.POST.get('employees_address')
#         employee_email = request.POST.get('employee_email')

#         # Create the User object
#         user = User.objects.create_user(username=username, password=password, first_name=first_name, last_name=last_name)

#         # Create the Employees object associated with the User
#         Employees.objects.create(
#             user=user,
#             employees_photo=employees_photo,
#             employees_job=employees_job,
#             employees_number=employees_number,
#             employees_address=employees_address,
#             employee_email=employee_email,
#         )
#         return redirect('employee-list')

#     return render(request, 'employee_create.html')

@login_required


def employee_create(request):
    if request.method == 'POST':
        username = request.POST.get('username')
        password = request.POST.get('password')
        first_name = request.POST.get('first_name')
        last_name = request.POST.get('last_name')
        employees_photo = request.FILES.get('employees_photo')
        employees_job = request.POST.get('employees_job')
        employees_number = request.POST.get('employees_number')
        employees_address = request.POST.get('employees_address')
        employee_email = request.POST.get('employee_email')

        try:
            # Create the User object
            user = User.objects.create_user(username=username, password=password, first_name=first_name, last_name=last_name)

            # Create the Employees object associated with the User
            Employees.objects.create(
                user=user,
                employees_photo=employees_photo,
                employees_job=employees_job,
                employees_number=employees_number,
                employees_address=employees_address,
                employee_email=employee_email,
            )

            return redirect('employee-list')

        except IntegrityError:
            # Handle the unique constraint failure
            messages.error(request, 'Username already exists. Please choose a different username.')
            return render(request, 'employee_create.html')

    return render(request, 'employee_create.html')



@login_required
def employee_update(request, pk):
    employee = Employees.objects.get(pk=pk)
    user = employee.user

    if request.method == 'POST':
        username = request.POST.get('username')
        password = request.POST.get('password')
        first_name = request.POST.get('first_name')
        last_name = request.POST.get('last_name')
        employees_photo = request.FILES.get('employees_photo')
        employees_job = request.POST.get('employees_job')
        employees_number = request.POST.get('employees_number')
        employees_address = request.POST.get('employees_address')
        employee_email = request.POST.get('employee_email')

        # Update the User object
        user.username = username
        if password:
            user.set_password(password)
        user.first_name = first_name
        user.last_name = last_name
        user.save()

        # Update the Employees object
        employee.employees_photo = employees_photo
        employee.employees_job = employees_job
        employee.employees_number = employees_number
        employee.employees_address = employees_address
        employee.employee_email = employee_email
        employee.save()

        return redirect('employee-list')

    return render(request, 'employee_update.html', {'employee': employee})


@login_required
def employee_delete(request, pk):
    employee = Employees.objects.get(pk=pk)
    user = employee.user  # Get the associated User object
    # Delete both the Employees and User objects
    employee.delete()
    user.delete()
    return redirect('employee-list')


# -------------------------------- Timesheet ---------------------------------------- #

@login_required
def timesheet_list(request):
    employee = request.user

    if employee.is_superuser:
        # Admin view: list all timesheets
        employees = Employees.objects.all()  # Fetch all employees for the dropdown
        projects = Projects.objects.all()  # Fetch all projects for the dropdown
        timesheets = TimeSheet.objects.all().order_by('-id')

        if request.method == 'GET':
            selected_employee = request.GET.get('employee')
            start_date = request.GET.get('start_date')
            end_date = request.GET.get('end_date')
            project_id = request.GET.get('project')
            status = request.GET.get('status')

            if selected_employee:
                timesheets = timesheets.filter(employee_id=selected_employee)
            if start_date:
                timesheets = timesheets.filter(date__gte=start_date)
            if end_date:
                timesheets = timesheets.filter(date__lte=end_date)
            if project_id:
                timesheets = timesheets.filter(project_id=project_id)
            if status:
                timesheets = timesheets.filter(status=status)

        for timesheet in timesheets:
            start_datetime = datetime.combine(datetime.now().date(), timesheet.start_time)
            end_datetime = datetime.combine(datetime.now().date(), timesheet.end_time)
            duration = end_datetime - start_datetime  # Calculate the duration as a timedelta object

            # Convert the timedelta to minutes
            total_minutes = int(duration.total_seconds() / 60)
            timesheet.duration = f"{total_minutes:02d}"



        if 'export' in request.GET:
            response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
            response['Content-Disposition'] = 'attachment; filename=timesheets.xlsx'

            workbook = Workbook()
            worksheet = workbook.active

            # Write headers
            headers = ['Employee', 'Date', 'Start Time', 'End Time', 'Project', 'Location', 'Type', 'Description', 'Status', 'Duration']
            for col_num, header in enumerate(headers, 1):
                worksheet.cell(row=1, column=col_num, value=header)

            # Write data
            for row_num, timesheet in enumerate(timesheets, 2):
                worksheet.cell(row=row_num, column=1, value=f"{timesheet.employee.first_name} {timesheet.employee.last_name}")
                column_letter = get_column_letter(1)  # Column B (index 2)
                worksheet.column_dimensions[column_letter].width = 15

                worksheet.cell(row=row_num, column=2, value=timesheet.date)
                # Set the column width for the date column
                column_letter = get_column_letter(2)  # Column B (index 2)
                worksheet.column_dimensions[column_letter].width = 15  # Adjust the width as needed

                worksheet.cell(row=row_num, column=3, value=timesheet.start_time)
                column_letter = get_column_letter(3)  # Column B (index 2)
                worksheet.column_dimensions[column_letter].width = 15

                worksheet.cell(row=row_num, column=4, value=timesheet.end_time)
                column_letter = get_column_letter(4)  # Column B (index 2)
                worksheet.column_dimensions[column_letter].width = 15

                worksheet.cell(row=row_num, column=5, value=timesheet.project.project_name)
                column_letter = get_column_letter(5)  # Column B (index 2)
                worksheet.column_dimensions[column_letter].width = 15

                worksheet.cell(row=row_num, column=6, value=timesheet.location)
                column_letter = get_column_letter(6)  # Column B (index 2)
                worksheet.column_dimensions[column_letter].width = 15

                worksheet.cell(row=row_num, column=7, value=timesheet.type)
                column_letter = get_column_letter(7)  # Column B (index 2)
                worksheet.column_dimensions[column_letter].width = 15

                worksheet.cell(row=row_num, column=8, value=timesheet.description)
                column_letter = get_column_letter(8)  # Column B (index 2)
                worksheet.column_dimensions[column_letter].width = 15

                worksheet.cell(row=row_num, column=9, value=timesheet.status)
                column_letter = get_column_letter(9)  # Column B (index 2)
                worksheet.column_dimensions[column_letter].width = 15

                worksheet.cell(row=row_num, column=10, value=timesheet.duration)
                column_letter = get_column_letter(10)  # Column B (index 2)
                worksheet.column_dimensions[column_letter].width = 15


            workbook.save(response)
            return response

        return render(request, 'timesheet_list.html', {'timesheets': timesheets, 'employees': employees, 'projects': projects})
    else:
        # Employee view: list timesheets based on status and date range
        start_date = request.GET.get('start_date')
        end_date = request.GET.get('end_date')

        approved_timesheets = TimeSheet.objects.filter(employee=employee, status='Approved').order_by('-id')
        pending_timesheets = TimeSheet.objects.filter(employee=employee, status='Pending').order_by('-id')
        rejected_timesheets = TimeSheet.objects.filter(employee=employee, status='Rejected').order_by('-id')

        if start_date:
            approved_timesheets = approved_timesheets.filter(date__gte=start_date)
            pending_timesheets = pending_timesheets.filter(date__gte=start_date)
            rejected_timesheets = rejected_timesheets.filter(date__gte=start_date)
        if end_date:
            approved_timesheets = approved_timesheets.filter(date__lte=end_date)
            pending_timesheets = pending_timesheets.filter(date__lte=end_date)
            rejected_timesheets = rejected_timesheets.filter(date__lte=end_date)

        for timesheet in pending_timesheets:
            start_datetime = datetime.combine(datetime.now().date(), timesheet.start_time)
            end_datetime = datetime.combine(datetime.now().date(), timesheet.end_time)
            duration = end_datetime - start_datetime  # Calculate the duration as a timedelta object

            # Convert the timedelta to string format
            hours, remainder = divmod(duration.total_seconds(), 3600)
            minutes, seconds = divmod(remainder, 60)
            timesheet.duration = f"{int(hours):02d}:{int(minutes):02d}:{int(seconds):02d}"

        for timesheet in rejected_timesheets:
            start_datetime = datetime.combine(datetime.now().date(), timesheet.start_time)
            end_datetime = datetime.combine(datetime.now().date(), timesheet.end_time)
            duration = end_datetime - start_datetime  # Calculate the duration as a timedelta object

            # Convert the timedelta to string format
            hours, remainder = divmod(duration.total_seconds(), 3600)
            minutes, seconds = divmod(remainder, 60)
            timesheet.duration = f"{int(hours):02d}:{int(minutes):02d}:{int(seconds):02d}"

        for timesheet in approved_timesheets:
            start_datetime = datetime.combine(datetime.now().date(), timesheet.start_time)
            end_datetime = datetime.combine(datetime.now().date(), timesheet.end_time)
            duration = end_datetime - start_datetime  # Calculate the duration as a timedelta object

            # Convert the timedelta to string format
            hours, remainder = divmod(duration.total_seconds(), 3600)
            minutes, seconds = divmod(remainder, 60)
            timesheet.duration = f"{int(hours):02d}:{int(minutes):02d}:{int(seconds):02d}"

        return render(request, 'timesheet_list.html', {'approved_timesheets': approved_timesheets,
                                                       'pending_timesheets': pending_timesheets,
                                                       'rejected_timesheets': rejected_timesheets})



@login_required  # Only logged-in employees can access these views
def create_timesheet(request):
    if request.method == 'POST':
        # Retrieve form data
        dates = request.POST.get('date', date.today().strftime('%Y-%m-%d'))
        start_time = request.POST['start_time']
        end_time = request.POST['end_time']
        location = request.POST['location']
        type = request.POST['type']
        type2 = request.POST['type2']
        project_id = request.POST['project']
        description = request.POST['description']
        
        # Create a new timesheet
        timesheet = TimeSheet.objects.create(
            employee=request.user,
            date=dates,
            start_time=start_time,
            end_time=end_time,
            location = location,
            type = type,
            type2 = type2,
            project_id=project_id,
            description = description
        )
        
        return redirect('timesheet-list')
    
    # Render the create timesheet form
    projects = Projects.objects.all()
    return render(request, 'create_timesheet.html', {'projects': projects, 'date': date.today()})



# @login_required  # Only logged-in admins can access this view
# def update_timesheet_status(request, pk):
#     if request.method == 'POST':
#         status = request.POST['status']
        
#         # Update timesheet status
#         timesheet = TimeSheet.objects.get(pk=pk)
#         timesheet.status = status
#         timesheet.save()
        
#         return redirect('timesheet-list')
    
#     # Render the update timesheet status form
#     timesheet = TimeSheet.objects.get(pk=pk)
#     return render(request, 'update_timesheet_status.html', {'timesheet': timesheet})


@login_required  # Only logged-in admins can access this view
def update_timesheet_status(request, pk):
    if request.method == 'POST':
        status = request.POST['status']
        
        # Update timesheet status
        timesheet = TimeSheet.objects.get(pk=pk)
        timesheet.status = status
        timesheet.save()
        
        return redirect('timesheet-list')
    
    # Calculate duration for the timesheet
    # timesheet_obj = TimeSheet.objects.get(pk=pk)
    # start_datetime = datetime.combine(datetime.now().date(), timesheet_obj.start_time)
    # end_datetime = datetime.combine(datetime.now().date(), timesheet_obj.end_time)
    # duration = end_datetime - start_datetime  # Calculate the duration as a timedelta object

    # # Convert the timedelta to string format
    # hours, remainder = divmod(duration.total_seconds(), 3600)
    # minutes, seconds = divmod(remainder, 60)
    # duration_str = f"{int(hours):02d}:{int(minutes):02d}:{int(seconds):02d}"
    
    timesheet_obj = TimeSheet.objects.get(pk=pk)
    start_datetime = datetime.combine(datetime.now().date(), timesheet_obj.start_time)
    end_datetime = datetime.combine(datetime.now().date(), timesheet_obj.end_time)
    duration = end_datetime - start_datetime  # Calculate the duration as a timedelta object

    # Convert the timedelta to total minutes
    total_minutes = duration.total_seconds() // 60
    duration_str = f"{int(total_minutes):02d}"

    # Render the update timesheet status form
    return render(request, 'update_timesheet_status.html', {'timesheet': timesheet_obj, 'duration': duration_str})



def delete_timesheet(request, pk):
    timesheet = TimeSheet.objects.get(pk=pk)
    timesheet.delete()
    return redirect('timesheet-list')






# -------------------------- profile ----------------------------- #


@login_required
def profile(request):
    employee = Employees.objects.get(user=request.user)
    return render(request, 'profile.html', {'employee': employee})

@login_required
def profileedit(request, pk):
    employee = Employees.objects.get(pk=pk)
    user = employee.user

    if request.method == 'POST':
        # username = request.POST.get('username')
        password = request.POST.get('password')
        first_name = request.POST.get('first_name')
        last_name = request.POST.get('last_name')
        
        
        employees_number = request.POST.get('employees_number')
        employees_address = request.POST.get('employees_address')
        employee_email = request.POST.get('employee_email')

        # Update the User object
        # user.username = username
        # if password:
        #     user.set_password(password)
        user.first_name = first_name
        user.last_name = last_name
        user.save()

        # Update the Employees object
        
       
        employee.employees_number = employees_number
        employee.employees_address = employees_address
        employee.employee_email = employee_email
        employee.save()

        return redirect('profile')
    return render(request, 'profileedit.html', {'employee': employee})


def profilephoto(request, pk):
    employee = Employees.objects.get(pk=pk)
    user = employee.user

    if request.method == 'POST':
        
        employees_photo = request.FILES.get('employees_photo')
        


        # Update the Employees object
        employee.employees_photo = employees_photo
        
        employee.save()

        return redirect('profile')
    return render(request, 'profilephoto.html', {'employee': employee})
